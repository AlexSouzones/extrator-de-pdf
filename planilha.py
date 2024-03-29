import openpyxl
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Alignment


planilha_caminho = ""
hora_data = datetime.now()
hora_data = hora_data.strftime("%H:%M")

def criar_planilha(caminho):
  global planilha_caminho
  wb = openpyxl.Workbook()

  sheet = wb.active
  sheet.column_dimensions['A'].width = 20
  sheet['A1'] = 'codigo'
  sheet.column_dimensions['B'].width = 30
  sheet['B1'] = 'projeto'
  sheet.column_dimensions['C'].width = 20
  sheet['C1'] = 'Quantidade de alunos'
  sheet.column_dimensions['D'].width = 20
  sheet['D1'] = 'Assessoramentos'
  sheet.column_dimensions['E'].width = 20
  sheet['E1'] = 'Quantidade eventos'
  sheet.column_dimensions['F'].width = 20
  sheet['F1'] = 'Grupos Beneficiados'
  sheet.column_dimensions['G'].width = 20
  sheet['G1'] = 'Publico Total'
  sheet.column_dimensions['H'].width = 20
  sheet['H1'] = 'Num Atendimentos'
  sheet.column_dimensions['I'].width = 20
  sheet['I1'] = 'Parcerias de Fora'
  sheet.column_dimensions['J'].width = 20
  sheet['J1'] = 'Empresas Entidades'
  sheet.column_dimensions['K'].width = 20
  sheet['K1'] = 'Parceria Setor Pub'
  sheet.column_dimensions['L'].width = 20
  sheet['L1'] = 'Terceiro Setor'
  sheet.column_dimensions['M'].width = 20
  sheet['M1'] = 'Material'
  sheet.column_dimensions['N'].width = 20
  sheet['N1'] = 'Arrecadações'


  for linha in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=14):
    for cell in linha:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

  nome_planilha = ""
  if os.path.isdir(caminho):
      arquivos = os.listdir(caminho)
      numero_arquivos = len(arquivos)+1
      nome_planilha = f'{caminho}/DADOS EXTRAÍDOS({numero_arquivos}).xlsx'
  planilha_caminho = nome_planilha
      
  wb.save(nome_planilha)
  


def preencher_planilha(codigo, projeto, qtd_aluno, assessoramentos, qtd_eventos, grupos_beneficiados, publico_total, numero_atendimentos,
                       parcerias_profi_fora, parcerias_empresas_entidades, parceria_setor_publico, parceria_terceiro_setor,
                       material_didatico, arrecadacao_donativos):
  planilha = openpyxl.load_workbook(planilha_caminho)
  nome_da_folha = 'Sheet'
  folha = planilha[nome_da_folha]

  valores = [codigo,projeto, qtd_aluno, assessoramentos, qtd_eventos, grupos_beneficiados, publico_total, numero_atendimentos,
             parcerias_profi_fora, parcerias_empresas_entidades, parceria_setor_publico, parceria_terceiro_setor,
             material_didatico, arrecadacao_donativos]

  proxima_linha_vazia = folha.max_row + 1

  for c in range(1, len(valores)+1):
      celula = folha.cell(row=proxima_linha_vazia, column=c)
      celula.value = valores[c-1]

  planilha.save(planilha_caminho)

  planilha.close()
